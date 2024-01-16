#!/bin/python3
from datetime import date as dt
from openpyxl import Workbook as wb
from openpyxl import font
from openpyxl import Side, Border, Alignment
import os
import subprocess

# Specify the absolute path for the Excel file
excel_file_path = '/home/guest/_repos/Scripting/Own_attempts/S2/excelexample.xlsx'
wb = wb()
wba = wb.active
################################################################################
#workflow#

wba['A1'] = 99

wba.append([1,2,3])

wba['A2'] = dt.today()

wba.title = "my first worksheet"

for row in range(3, 10):
    wba.append(range(5))

for row in range(11,100):
    wba.append(range(100))

wba2 = wb.create_sheet(title="WS2")
#wba2['B2'] = 'This is B2'

for row in range(1,11):
    for col in range(1,11):
        var = "C" + str(col) + "R" + str(row)
        _ = wba2.cell(column=col, row=row, value="{}".format(var))

wba2.merge_cells('A2:D2')
wba2.merge_cells(start_row=3, start_column=1, end_row=5, end_column=4)
wba3 = wb.create_sheet(title="WS3")

styledcell = wba3["A1"]
styledcell.font = font(name='Arial', size=20, color="FF0000")
styledcell2 = wba3['A2']
styledcell2.font = font(name='Tahoma', size=16, color="0000FF", bold=True)

styling = font(name='Tahoma', size=16, color="0000FF", bold=True)

a5 = wba3['A5']
a5.font = styling
d10 = wba3['D10']
d10.value = "borders"
d10.font = font(name='Tahoma', size=10, color="600060", bold=True)
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
d10.border = Border(top=double, left=thin, right=thin, bottom=double)
d10.alignment = Alignment(horizontal="center", vertical="center")
# after written to cell in column 
# you can modify column dimensions
wba3.column_dimensions["C"].width = 30.0
wba3.row_dimensions[10].height = 40.0


################################################################################
wb.save(excel_file_path)
subprocess.run(['xdg-open', excel_file_path])